
import openpyxl
from openpyxl.styles import Font, Alignment
from typing import Dict
import os
from datetime import datetime

def export_to_excel(processed_data: Dict) -> str:
    """Generate an Excel file from the processed data."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis Result"

    # --- Styles ---
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    
    # --- Helpers ---
    def add_section(title, start_row):
        ws.cell(row=start_row, column=1, value=title).font = section_font
        return start_row + 1

    # --- Content ---
    current_row = 1
    ws.cell(row=current_row, column=1, value="帳目分析結果").font = header_font
    current_row += 2

    # --- Summary Section ---
    current_row = add_section("總覽", current_row)
    summary_data = processed_data.get('extracted_data', {})
    custom_fields = summary_data.get('custom_fields', {})

    ws.cell(row=current_row, column=1, value="日期")
    ws.cell(row=current_row, column=2, value=summary_data.get('date', 'N/A'))
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="總收入 (折扣前)")
    ws.cell(row=current_row, column=2, value=summary_data.get('calculated_total'))
    current_row += 1

    if custom_fields:
        ws.cell(row=current_row, column=1, value="總折扣")
        ws.cell(row=current_row, column=2, value=custom_fields.get('total_discount'))
        current_row += 1
        ws.cell(row=current_row, column=1, value="最終結餘")
        ws.cell(row=current_row, column=2, value=custom_fields.get('final_balance'))
        current_row += 1

    current_row += 1 # Spacer

    # --- Items Section ---
    current_row = add_section("項目明細", current_row)
    items = summary_data.get('items', [])
    
    # Header
    ws.cell(row=current_row, column=1, value="項目").font = section_font
    ws.cell(row=current_row, column=2, value="金額").font = section_font
    ws.cell(row=current_row, column=3, value="折扣").font = section_font
    ws.cell(row=current_row, column=4, value="備註").font = section_font
    current_row += 1

    for item in items:
        ws.cell(row=current_row, column=1, value=item.get('name', item.get('item')))
        ws.cell(row=current_row, column=2, value=item.get('amount', item.get('cost')))
        ws.cell(row=current_row, column=3, value=item.get('discount', 0))
        ws.cell(row=current_row, column=4, value=item.get('note', ''))
        current_row += 1

    # --- Auto-fit columns ---
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # --- Save File ---
    export_dir = "data/exports"
    os.makedirs(export_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(export_dir, f"analysis_{timestamp}.xlsx")
    
    wb.save(file_path)
    
    return file_path
